import pandas as pd
import webbrowser
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os
from typing import Dict, List, Optional
import urllib.parse
import re
import io

class WhatsAppOrderProcessor:
    def __init__(self, config: Optional[Dict] = None):
        """
        Initialize with configuration for column mappings.
        """
        self.config = config or {
            'phone_col': 'Phone (Billing)',
            'name_col': 'Full Name (Billing)',
            'order_id_col': 'Order ID',
            'product_col': 'Product Name (main)',
            'sku_col': 'SKU',
            'quantity_col': 'Quantity',
            'price_col': 'Order Line Subtotal',
            'payment_method_col': 'Payment Method Title',
            'address_col': 'Address 1&2 (Billing)',
            'order_total_col': 'Order Total Amount',
            'city_col': 'City, State, Zip (Billing)'
        }
        self.column_map = {}

    def clean_phone_number(self, phone: str) -> str:
        """Clean and format phone number for WhatsApp."""
        if pd.isna(phone):
            return ""
        # Remove any non-digit characters
        phone = ''.join(filter(str.isdigit, str(phone)))
        # Ensure it has country code
        if phone.startswith('0'):
            phone = '0' + phone[1:]  # Assuming BD numbers
        elif not phone.startswith(('88', '+88')):
            phone = '0' + phone
        return phone
        
    def format_text(self, text: str) -> str:
        """
        Cleans and formats a text string for names or addresses.
        - Handles names with periods (e.g., "Ad.suman" -> "Ad. Suman")
        - Handles addresses with commas (e.g., "islambag,panchagarh" -> "Islambag, Panchagarh")
        - Handles multiple spaces around commas (e.g., "Prodhan Para , Gobindoganj" -> "Prodhan Para, Gobindoganj")
        - Capitalizes first letter of each word
        - Preserves special characters like hyphens and apostrophes
        """
        if not isinstance(text, str) or not text.strip():
            return ""
            
        # First, handle the entire text as a whole for address parts
        if ',' in text:
            # Split by comma and process each part
            parts = [p.strip() for p in text.split(',')]
            formatted_parts = []
            
            for part in parts:
                if not part:
                    continue
                    
                # Process words within each part
                words = part.split()
                formatted_words = []
                
                for word in words:
                    # Handle hyphenated words
                    if '-' in word:
                        subparts = word.split('-')
                        subparts = [sp[0].upper() + sp[1:].lower() for sp in subparts if sp]
                        word = '-'.join(subparts)
                    # Handle words with apostrophes
                    elif "'" in word:
                        subparts = word.split("'")
                        subparts = [sp[0].upper() + sp[1:].lower() for sp in subparts if sp]
                        word = "'".join(subparts)
                    # Handle words with periods
                    elif '.' in word and not word.endswith('.'):
                        subparts = word.split('.')
                        subparts = [sp[0].upper() + sp[1:].lower() for sp in subparts if sp]
                        word = '. '.join(subparts)
                    # Regular words
                    else:
                        word = word[0].upper() + word[1:].lower()
                    
                    formatted_words.append(word)
                
                formatted_parts.append(' '.join(formatted_words))
            
            # Join with comma + space
            return ', '.join(formatted_parts)
            
        # For text without commas, handle as before
        else:
            # Handle period-separated parts (e.g., "Ad.suman" -> "Ad. Suman")
            text = re.sub(r'\.(\w)', r'. \1', text.strip())
            
            # Split into words
            words = text.split()
            formatted_words = []
            
            for word in words:
                if not word.strip():
                    continue
                    
                # Handle hyphenated words
                if '-' in word:
                    parts = word.split('-')
                    parts = [p[0].upper() + p[1:].lower() for p in parts if p]
                    word = '-'.join(parts)
                # Handle apostrophes
                elif "'" in word:
                    parts = word.split("'")
                    parts = [p[0].upper() + p[1:].lower() for p in parts if p]
                    word = "'".join(parts)
                # Handle words with periods
                elif '.' in word and not word.endswith('.'):
                    parts = word.split('.')
                    parts = [p[0].upper() + p[1:].lower() for p in parts if p]
                    word = '. '.join(parts)
                # Regular words
                else:
                    word = word[0].upper() + word[1:].lower()
                
                formatted_words.append(word)
            
            return ' '.join(formatted_words)

    def detect_gender_salutation(self, name: str) -> str:
        """
        Detect gender from name and return appropriate salutation.
        Default to 'Sir'.
        """
        if not isinstance(name, str):
            return "Sir"
            
        name_lower = name.lower()
        
        # Common female titles/names in BD/South Asian context
        female_indicators = [
            'ms', 'miss', 'mrs', 'mst', 'begum', 'khatun', 'akter', 'parvin', 
            'sultana', 'jahan', 'bibi', 'rani', 'devi', 'nahar', 'ferdous', 
            'ara', 'banu', 'fatema', 'aisha', 'khadija', 'nusrat', 'farhana', 
            'sadia', 'jannatul', 'sumaiya', 'tanjina', 'fariha', 'sharmin',
            'nasrin', 'salma', 'shirin', 'rumana', 'sabina'
        ]
        
        # Check for exact matches or parts
        parts = name_lower.replace('.', ' ').split()
        for part in parts:
            if part in female_indicators:
                return "Madam"
                
        return "Sir"

    def format_name(self, name):
        """Format name with proper capitalization and spacing."""
        if pd.isna(name):
            return ""
        return self.format_text(str(name))

    def format_address(self, *address_parts):
        """Format address parts with proper capitalization and spacing."""
        formatted_parts = []
        for part in address_parts:
            if pd.isna(part) or str(part).strip() == '':
                continue
            formatted_part = self.format_text(str(part))
            if formatted_part:  # Only add non-empty parts
                formatted_parts.append(formatted_part)
        return '\n'.join(formatted_parts)

    def process_orders(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Process and group orders by phone number.
        """
        # Clean and standardize column names
        df.columns = [str(col).strip() for col in df.columns]
        
        # Map columns to standard names
        self.column_map = {k: v for k, v in self.config.items() if v in df.columns}
        
        # Check for required columns
        required = ['phone_col', 'name_col', 'product_col']
        missing = [self.config[col] for col in required if self.config[col] not in df.columns]
        if missing:
            raise ValueError(f"Required columns not found: {missing}")

        # Clean phone numbers and format names/addresses
        phone_col = self.config['phone_col']
        name_col = self.config['name_col']
        address_col = self.config['address_col']
        
        df[phone_col] = df[phone_col].apply(self.clean_phone_number)
        df[name_col] = df[name_col].apply(self.format_name)
        
        # Format address if address column exists
        if address_col in df.columns:
            df[address_col] = df[address_col].apply(lambda x: self.format_text(str(x)) if pd.notna(x) else "")
        
        # Format other address components if they exist
        for key in ['city_col']:
            col = self.config.get(key)
            if col and col in df.columns:
                df[col] = df[col].apply(lambda x: self.format_text(str(x)) if pd.notna(x) else "")
        
        # Group by phone number and aggregate
        agg_funcs = {
            name_col: 'first',
            address_col: 'first',
            self.config['order_id_col']: lambda x: ', '.join(map(str, x.unique())),
            self.config['product_col']: lambda x: '\n- '.join(x),
            self.config['quantity_col']: lambda x: '\n- '.join(map(str, x)),
            self.config['price_col']: lambda x: '\n- '.join(map(str, x)),
            self.config['order_total_col']: 'sum'
        }
        
        # Add any additional address columns to aggregation
        for key in ['city_col', 'payment_method_col']:
            col = self.config.get(key)
            if col and col in df.columns:
                agg_funcs[col] = 'first'
        
        # Integrate SKU into Product Name if SKU column exists
        product_col = self.config['product_col']
        sku_col = self.config.get('sku_col')
        
        if sku_col and sku_col in df.columns:
            # Create a combined column for aggregation "Item Name - SKU"
            # We do this on the original df before groupby
            def combine_product_sku(row):
                p = str(row[product_col]) if pd.notna(row[product_col]) else ""
                s = str(row[sku_col]) if pd.notna(row[sku_col]) else ""
                
                if s.strip() and s.lower() != 'nan':
                    return f"{p} - {s}"
                return p
            
            df[product_col] = df.apply(combine_product_sku, axis=1)

        result_df = df.groupby(phone_col, as_index=False).agg(agg_funcs)
        
        # Ensure all address components are properly formatted in the final output
        if address_col in result_df.columns:
            result_df[address_col] = result_df[address_col].apply(
                lambda x: self.format_text(str(x)) if pd.notna(x) else ""
            )
            
        return result_df

    def create_whatsapp_links(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Create WhatsApp links for each order group.
        """
        df['whatsapp_link'] = None
        phone_col = self.config['phone_col']
        
        for idx, row in df.iterrows():
            phone = row[phone_col]
            if not phone:
                continue
                
            # Format name and address
            name = self.format_name(row[self.config['name_col']])
            formatted_address = self.format_address(
                row.get(self.config['address_col'], ''),
                row.get(self.config.get('city_col', 'City, State, Zip (Billing)'), '')
            )
            
            # Format phone number for display
            phone_display = f"+880{row[phone_col][-10:]}" if len(str(row[phone_col])) >= 10 else f"+88{row[phone_col]}"

            # Determine Salutation
            salutation = self.detect_gender_salutation(row[self.config['name_col']])

            # Build message
            message = f"*Order Verification From DEEN Commerce*\n\n"
            message += f"Assalamu Alaikum, {salutation}!\n\n"
            message += f"Dear {name},\n\n"
            message += "Please verify your order details:\n\n"
            message += f"*Order ID:* {row[self.config['order_id_col']]}\n\n"
            
            # Add order items
            message += "*Your Order:*\n"
            products = str(row[self.config['product_col']]).split('\n- ')
            # sizes = str(row[self.config['size_col']]).split('\n- ') # Size skipped as requested
            quantities = str(row[self.config['quantity_col']]).split('\n- ')
            prices = str(row[self.config['price_col']]).split('\n- ')
            
            for i in range(len(products)):
                if i < len(products):
                    msg_line = f"- {products[i].strip()}"
                    # Size integration handled in product name or skipped
                    if i < len(quantities):
                        msg_line += f" - Qty: {quantities[i].strip()}"
                    if i < len(prices):
                        msg_line += f" - Price: {prices[i].strip()} BDT"
                    message += msg_line + "\n"
            
            # Calculate Collectable Amount
            total_amount = float(row[self.config['order_total_col']])
            collectable_amount = total_amount
            
            payment_col = self.config.get('payment_method_col')
            if payment_col and payment_col in row and pd.notna(row[payment_col]):
                payment_method = str(row[payment_col]).lower()
                if any(x in payment_method for x in ['bkash', 'online', 'ssl', 'paid']):
                    collectable_amount = 0

            # Add total and address
            if collectable_amount == 0:
                 message += f"\n*Total Amount:* {total_amount:.2f} BDT (PAID)\n"
                 message += f"*Collectable Amount:* 0 BDT\n\n"
            else:
                 message += f"\n*Total Amount:* {total_amount:.2f} BDT\n\n"

            message += f"*Shipping Address:*\n{formatted_address}\n\n"
            # Contact Number skipped as requested
            
            message += "Please confirm your order.\n"
            message += "If any correction needed, let us know the possible adjustment.\n\n"
            message += "*Delivery fees apply for returns.*\n\n"
            message += "Thank you for shopping with DEEN Commerce! https://deencommerce.com/products/sale/"
                        
            # Create WhatsApp URL
            encoded_message = urllib.parse.quote(message)
            df.at[idx, 'whatsapp_link'] = f"https://wa.me/+88{phone}?text={encoded_message}"
        
        return df

    def process_excel(self, input_file: str, output_file: Optional[str] = None) -> str:
        """
        Process the Excel file and create WhatsApp-enabled version.
        """
        try:
            # Read input file
            df = pd.read_excel(input_file)
            
            # Process orders
            processed_df = self.process_orders(df)
            
            # Create WhatsApp links
            result_df = self.create_whatsapp_links(processed_df)
            
            # Generate output filename
            if not output_file:
                base, ext = os.path.splitext(input_file)
                output_file = f"{base}_whatsapp{ext}"
            
            # Save with formatting
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Orders')
                
                # Formatting
                workbook = writer.book
                worksheet = writer.sheets['Orders']
                
                # Style header
                header_fill = PatternFill(
                    start_color="4F81BD",
                    end_color="4F81BD",
                    fill_type="solid"
                )
                header_font = Font(color="FFFFFF", bold=True)
                
                for col in range(1, len(result_df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Format WhatsApp links
                whatsapp_col = result_df.columns.get_loc('whatsapp_link') + 1
                for row in range(2, len(result_df) + 2):
                    cell = worksheet.cell(row=row, column=whatsapp_col)
                    cell.hyperlink = cell.value
                    cell.value = 'Send WhatsApp'
                    cell.font = Font(color='0000FF', underline='single')
                    cell.alignment = Alignment(horizontal='center')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
            
            print(f"Successfully created WhatsApp-enabled Excel file: {output_file}")
            return output_file
            
        except Exception as e:
            print(f"An error occurred: {e}")
            raise

    def generate_excel_bytes(self, df: pd.DataFrame) -> bytes:
        """
        Generate Excel file bytes with formatting from the dataframe.
        """
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Orders')
            
            # Formatting
            workbook = writer.book
            worksheet = writer.sheets['Orders']
            
            # Style header
            header_fill = PatternFill(
                start_color="4F81BD",
                end_color="4F81BD",
                fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Format WhatsApp links
            if 'whatsapp_link' in df.columns:
                whatsapp_col = df.columns.get_loc('whatsapp_link') + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=whatsapp_col)
                    cell.hyperlink = cell.value
                    cell.value = 'Send WhatsApp'
                    cell.font = Font(color='0000FF', underline='single')
                    cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        return output.getvalue()

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Process Excel file and add WhatsApp integration.')
    parser.add_argument('input_file', help='Path to the input Excel file')
    parser.add_argument('-o', '--output', help='Path to save the output file (optional)')
    
    args = parser.parse_args()
    
    processor = WhatsAppOrderProcessor()
    processor.process_excel(args.input_file, args.output)

if __name__ == "__main__":
    main()