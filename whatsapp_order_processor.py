import pandas as pd
import webbrowser
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os
from typing import Dict, List, Optional
import urllib.parse
import re

class WhatsAppOrderProcessor:
    def __init__(self, config: Optional[Dict] = None):
        """
        Initialize with configuration for column mappings.
        """
        self.config = config or {
            'phone_col': 'Phone (Billing)',
            'name_col': 'Full Name (Billing)',
            'order_id_col': 'Order ID',
            'product_col': 'Product Name (main)',
            'size_col': 'Product Variation',
            'quantity_col': 'Quantity',
            'price_col': 'Order Line Subtotal',
            'address_col': 'Address 1&2 (Billing)',
            'order_total_col': 'Order Total Amount'
        }
        self.column_map = {}

    def clean_phone_number(self, phone: str) -> str:
        """Clean and format phone number for WhatsApp."""
        if pd.isna(phone):
            return ""
        # Remove any non-digit characters
        phone = ''.join(filter(str.isdigit, str(phone)))
        # Ensure it has country code
        if phone.startswith('0'):
            phone = '0' + phone[1:]  # Assuming BD numbers
        elif not phone.startswith(('88', '+88')):
            phone = '0' + phone
        return phone
        
    def format_text(self, text: str) -> str:
        """
        Cleans and formats a text string for names or addresses.
        - Handles names with periods (e.g., "Ad.suman" -> "Ad. Suman")
        - Handles addresses with commas (e.g., "islambag,panchagarh" -> "Islambag, Panchagarh")
        - Handles multiple spaces around commas (e.g., "Prodhan Para , Gobindoganj" -> "Prodhan Para, Gobindoganj")
        - Capitalizes first letter of each word
        - Preserves special characters like hyphens and apostrophes
        """
        if not isinstance(text, str) or not text.strip():
            return ""
            
        # Handle the entire text as a whole for address parts
        if ',' in text:
            # Split by comma and process each part
            parts = [p.strip() for p in text.split(',')]
            formatted_parts = []
            
            for part in parts:
                if not part:
                    continue
                    
                # Process words within each part
                words = part.split()
                formatted_words = []
                
                for word in words:
                    # Handle hyphenated words
                    if '-' in word:
                        subparts = word.split('-')
                        subparts = [sp[0].upper() + sp[1:].lower() for sp in subparts if sp]
                        word = '-'.join(subparts)
                    # Handle words with apostrophes
                    elif "'" in word:
                        subparts = word.split("'")
                        subparts = [sp[0].upper() + sp[1:].lower() for sp in subparts if sp]
                        word = "'".join(subparts)
                    # Handle words with periods
                    elif '.' in word and not word.endswith('.'):
                        subparts = word.split('.')
                        subparts = [sp[0].upper() + sp[1:].lower() for sp in subparts if sp]
                        word = '. '.join(subparts)
                    # Regular words
                    else:
                        word = word[0].upper() + word[1:].lower()
                    
                    formatted_words.append(word)
                
                formatted_parts.append(' '.join(formatted_words))
            
            # Join with comma + space
            return ', '.join(formatted_parts)
            
        # For text without commas, handle as before
        else:
            # Handle period-separated parts (e.g., "Ad.suman" -> "Ad. Suman")
            text = re.sub(r'\.(\w)', r'. \1', text.strip())
            
            # Split into words
            words = text.split()
            formatted_words = []
            
            for word in words:
                if not word.strip():
                    continue
                    
                # Handle hyphenated words
                if '-' in word:
                    parts = word.split('-')
                    parts = [p[0].upper() + p[1:].lower() for p in parts if p]
                    word = '-'.join(parts)
                # Handle apostrophes
                elif "'" in word:
                    parts = word.split("'")
                    parts = [p[0].upper() + p[1:].lower() for p in parts if p]
                    word = "'".join(parts)
                # Handle words with periods
                elif '.' in word and not word.endswith('.'):
                    parts = word.split('.')
                    parts = [p[0].upper() + p[1:].lower() for p in parts if p]
                    word = '. '.join(parts)
                # Regular words
                else:
                    word = word[0].upper() + word[1:].lower()
                
                formatted_words.append(word)
            
            return ' '.join(formatted_words)

    def format_name(self, name):
        """Format name with proper capitalization and spacing."""
        if pd.isna(name):
            return ""
        return self.format_text(str(name))

    def format_address(self, *address_parts):
        """Format address parts with proper capitalization and spacing, removing redundancy but preserving important commas."""
        # BD district mapping
        bd_districts = {
            'bd-01': 'Bandarban',
            'bd-02': 'Barguna',
            'bd-03': 'Bogura',
            'bd-04': 'Brahmanbaria',
            'bd-05': 'Bagerhat',
            'bd-06': 'Barishal',
            'bd-07': 'Bhola',
            'bd-08': 'Cumilla',
            'bd-09': 'Chandpur',
            'bd-10': 'Chattogram',
            'bd-11': 'Cox\'s Bazar',
            'bd-12': 'Chuadanga',
            'bd-13': 'Dhaka',
            'bd-14': 'Dinajpur',
            'bd-15': 'Faridpur',
            'bd-16': 'Feni',
            'bd-17': 'Gopalganj',
            'bd-18': 'Gazipur',
            'bd-19': 'Gaibandha',
            'bd-20': 'Habiganj',
            'bd-21': 'Jamalpur',
            'bd-22': 'Jessore',
            'bd-23': 'Jhenaidah',
            'bd-24': 'Joypurhat',
            'bd-25': 'Jhalokati',
            'bd-26': 'Kishoreganj',
            'bd-27': 'Khulna',
            'bd-28': 'Kurigram',
            'bd-29': 'Khagrachhari',
            'bd-30': 'Kushtia',
            'bd-31': 'Lakshmipur',
            'bd-32': 'Lalmonirhat',
            'bd-33': 'Manikganj',
            'bd-34': 'Mymensingh',
            'bd-35': 'Munshiganj',
            'bd-36': 'Madaripur',
            'bd-37': 'Magura',
            'bd-38': 'Moulvibazar',
            'bd-39': 'Meherpur',
            'bd-40': 'Narayanganj',
            'bd-41': 'Netrakona',
            'bd-42': 'Narsingdi',
            'bd-43': 'Narail',
            'bd-44': 'Natore',
            'bd-45': 'Nawabganj',
            'bd-46': 'Nilphamari',
            'bd-47': 'Noakhali',
            'bd-48': 'Naogaon',
            'bd-49': 'Pabna',
            'bd-50': 'Pirojpur',
            'bd-51': 'Patuakhali',
            'bd-52': 'Panchagarh',
            'bd-53': 'Rajbari',
            'bd-54': 'Rajshahi',
            'bd-55': 'Rangpur',
            'bd-56': 'Rangamati',
            'bd-57': 'Shariatpur',
            'bd-58': 'Satkhira',
            'bd-59': 'Sirajganj',
            'bd-60': 'Sherpur',
            'bd-61': 'Sunamganj',
            'bd-62': 'Sylhet',
            'bd-63': 'Tangail',
            'bd-64': 'Thakurgaon'
        }
        
        # Process each part to replace BD-* and format text
        processed_parts = []
        for part in address_parts:
            if pd.isna(part) or not str(part).strip():
                continue
                
            part_str = str(part).strip()
            # Replace BD-* with district name
            for bd_code, district in bd_districts.items():
                part_str = part_str.replace(bd_code, district)
                part_str = part_str.replace(bd_code.upper(), district)
                part_str = part_str.replace(bd_code.lower(), district)
            
            # Format the text (preserving commas)
            formatted = self.format_text(part_str)
            if formatted:
                processed_parts.append(formatted)
        
        if not processed_parts:
            return ""
        
        # Join all parts with comma and space
        full_address = ', '.join(processed_parts)
        
        # Remove duplicate words while preserving comma structure
        words = re.split(r'([,])', full_address)
        seen_words = set()
        result = []
        
        i = 0
        while i < len(words):
            word = words[i].strip()
            
            # If it's a comma, add it to the result
            if word == ',':
                result.append(word)
                i += 1
                continue
                
            # Skip empty words
            if not word:
                i += 1
                continue
                
            # Check if we've seen this word before (case-insensitive)
            word_lower = word.lower()
            if word_lower not in seen_words:
                seen_words.add(word_lower)
                result.append(word)
                # If next token is a comma, add a space
                if i + 1 < len(words) and words[i+1] == ',':
                    result.append(' ')
            else:
                # If we're skipping a word that's followed by a comma, skip the comma too
                if i + 1 < len(words) and words[i+1] == ',':
                    i += 1
                    
            i += 1
        
        # Join the result and clean up any double spaces
        final_address = ''.join(result)
        final_address = re.sub(r'\s+', ' ', final_address).strip()
        final_address = re.sub(r',\s*,', ', ', final_address)  # Clean up double commas
        final_address = re.sub(r'\s*,\s*', ', ', final_address)  # Normalize spaces around commas
        
        return final_address  

    def process_orders(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Process and group orders by phone number.
        """
        # Clean and standardize column names
        df.columns = [str(col).strip() for col in df.columns]
        
        # Map columns to standard names
        self.column_map = {k: v for k, v in self.config.items() if v in df.columns}
        
        # Check for required columns
        required = ['phone_col', 'name_col', 'product_col']
        missing = [self.config[col] for col in required if self.config[col] not in df.columns]
        if missing:
            raise ValueError(f"Required columns not found: {missing}")

        # Clean phone numbers and format names
        phone_col = self.config['phone_col']
        name_col = self.config['name_col']
        address_col = self.config['address_col']
        
        df[phone_col] = df[phone_col].apply(self.clean_phone_number)
        df[name_col] = df[name_col].apply(self.format_name)
        
        # Combine address parts into a single formatted address
        if address_col in df.columns:
            # Create a combined address field
            df[address_col] = df.apply(
                lambda row: self.format_address(
                    row[address_col],
                    row.get('City, State, Zip (Billing)', '')
                ),
                axis=1
            )
        
        # Group by phone number and aggregate
        agg_funcs = {
            name_col: 'first',
            address_col: 'first',
            self.config['order_id_col']: lambda x: ', '.join(map(str, x.unique())),
            self.config['product_col']: lambda x: '\n- '.join(x),
            self.config['size_col']: lambda x: '\n- '.join(map(str, x)),
            self.config['quantity_col']: lambda x: '\n- '.join(map(str, x)),
            self.config['price_col']: lambda x: '\n- '.join(map(str, x)),
            self.config['order_total_col']: 'first'
        }
        
        result_df = df.groupby(phone_col, as_index=False).agg(agg_funcs)
        
        return result_df

    def create_whatsapp_links(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Create WhatsApp links for each order group.
        """
        df['whatsapp_link'] = None
        phone_col = self.config['phone_col']
        
        for idx, row in df.iterrows():
            phone = row[phone_col]
            if not phone:
                continue
                
            # Format name
            name = self.format_name(row[self.config['name_col']])
            
            # Format phone number for display
            phone_display = f"+880{row[phone_col][-10:]}" if len(str(row[phone_col])) >= 10 else f"+88{row[phone_col]}"

            # Build message
            message = f"*Order Verification From DEEN Commerce*\n\n"
            message += f"Assalamu Alaikum, Sir!\n\n"
            message += f"Dear {name},\n\n"
            message += "Please verify your order details:\n\n"
            message += f"*Order ID:* {row[self.config['order_id_col']]}\n\n"
            
            # Add order items
            message += "*Your Order:*\n"
            products = str(row[self.config['product_col']]).split('\n- ')
            sizes = str(row[self.config['size_col']]).split('\n- ')
            quantities = str(row[self.config['quantity_col']]).split('\n- ')
            prices = str(row[self.config['price_col']]).split('\n- ')
            
            for i in range(len(products)):
                if i < len(products):
                    msg_line = f"- {products[i].strip()}"
                    if i < len(sizes) and str(sizes[i]).strip() != 'nan':
                        msg_line += f" (Size: {sizes[i].strip()})"
                    if i < len(quantities):
                        msg_line += f" - Qty: {quantities[i].strip()}"
                    if i < len(prices):
                        msg_line += f" - Price: {prices[i].strip()} BDT"
                    message += msg_line + "\n"
            
            # Add total and address
            message += f"\n*Total Amount:* {float(row[self.config['order_total_col']]):.2f} BDT\n\n"
            message += f"*Shipping Address:*\n{row[self.config['address_col']]}\n\n"
            message += f"*Contact Number:* {phone_display}\n\n"
            message += "Please reply with 'YES or Confirmed' to confirm your order or 'NO' if there's any mistake and 'Cancel' to cancel the order or if you want to add or change any product plese inform us.\n"
            message += "Thank you for shopping with DEEN Commerce!"
                        
            # Create WhatsApp URL
            encoded_message = urllib.parse.quote(message)
            df.at[idx, 'whatsapp_link'] = f"https://wa.me/+88{phone}?text={encoded_message}"
        
        return df

    def process_excel(self, input_file: str, output_file: Optional[str] = None) -> str:
        """
        Process the Excel file and create WhatsApp-enabled version.
        """
        try:
            # Read input file
            df = pd.read_excel(input_file)
            
            # Process orders
            processed_df = self.process_orders(df)
            
            # Create WhatsApp links
            result_df = self.create_whatsapp_links(processed_df)
            
            # Generate output filename
            if not output_file:
                base, ext = os.path.splitext(input_file)
                output_file = f"{base}_whatsapp{ext}"
            
            # Save with formatting
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Orders')
                
                # Formatting
                workbook = writer.book
                worksheet = writer.sheets['Orders']
                
                # Style header
                header_fill = PatternFill(
                    start_color="4F81BD",
                    end_color="4F81BD",
                    fill_type="solid"
                )
                header_font = Font(color="FFFFFF", bold=True)
                
                for col in range(1, len(result_df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Format WhatsApp links
                whatsapp_col = result_df.columns.get_loc('whatsapp_link') + 1
                for row in range(2, len(result_df) + 2):
                    cell = worksheet.cell(row=row, column=whatsapp_col)
                    cell.hyperlink = cell.value
                    cell.value = 'Send WhatsApp'
                    cell.font = Font(color='0000FF', underline='single')
                    cell.alignment = Alignment(horizontal='center')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
            
            print(f"Successfully created WhatsApp-enabled Excel file: {output_file}")
            return output_file
            
        except Exception as e:
            print(f"An error occurred: {e}")
            raise

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Process Excel file and create WhatsApp links.')
    parser.add_argument('input_file', help='Input Excel file path')
    parser.add_argument('-o', '--output', help='Output Excel file path (optional)')
    
    args = parser.parse_args()
    
    processor = WhatsAppOrderProcessor()
    processor.process_excel(args.input_file, args.output)