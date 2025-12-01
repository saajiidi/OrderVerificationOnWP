import pandas as pd
import webbrowser
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
import os

def create_whatsapp_template(input_file, output_file=None):
    """
    Create an Excel file with WhatsApp sending functionality
    
    Args:
        input_file (str): Path to the input Excel file
        output_file (str, optional): Path to save the output file. 
                                   If None, will add '_whatsapp' to input filename.
    """
    try:
        # Read the input Excel file
        df = pd.read_excel(input_file)
        
        # Clean column names (remove extra spaces and make lowercase)
        df.columns = [str(col).strip().lower() for col in df.columns]
        
        # Generate output filename if not provided
        if output_file is None:
            base, ext = os.path.splitext(input_file)
            output_file = f"{base}_whatsapp{ext}"
        
        # Create a copy of the DataFrame for processing
        processed_df = df.copy()
        
        # Add WhatsApp action column
        processed_df['whatsapp_action'] = 'Send WhatsApp'
        
        # Reorder columns to put WhatsApp action first
        cols = ['whatsapp_action'] + [col for col in processed_df.columns if col != 'whatsapp_action']
        processed_df = processed_df[cols]
        
        # Save to Excel
        processed_df.to_excel(output_file, index=False, sheet_name='Orders')
        
        # Load the workbook to format it
        book = load_workbook(output_file)
        sheet = book.active
        
        # Format the header
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Format the WhatsApp action column
        whatsapp_col = 1  # First column
        for row in range(2, len(processed_df) + 2):
            # Create hyperlink for WhatsApp
            phone_cell = sheet.cell(row=row, column=cols.index('phone') + 1)
            phone = str(phone_cell.value).strip()
            
            # Format phone number (remove any non-digit characters)
            phone = ''.join(filter(str.isdigit, phone))
            
            # Create WhatsApp URL
            whatsapp_url = f"https://wa.me/{phone}?text="
            
            # Add message template
            message = "*Order Verification*\n\n"
            message += f"Dear {sheet.cell(row=row, column=cols.index('customer_name') + 1).value},\n\n"
            message += "Please verify your order details:\n\n"
            
            # Add order details
            message += f"Order ID: {sheet.cell(row=row, column=cols.index('order_id') + 1).value}\n"
            message += f"Product: {sheet.cell(row=row, column=cols.index('product_name') + 1).value}\n"
            message += f"Size: {sheet.cell(row=row, column=cols.index('size') + 1).value}\n"
            message += f"Quantity: {sheet.cell(row=row, column=cols.index('quantity') + 1).value}\n"
            message += f"Total: {sheet.cell(row=row, column=cols.index('total') + 1).value}\n\n"
            message += "*Shipping Address:*\n"
            message += f"{sheet.cell(row=row, column=cols.index('address') + 1).value}\n\n"
            message += "Please reply with 'YES' to confirm your order or 'NO' if there's any mistake.\n"
            message += "Thank you for shopping with us!"
            
            # URL encode the message
            import urllib.parse
            encoded_message = urllib.parse.quote(message)
            whatsapp_url += encoded_message
            
            # Create hyperlink in the cell
            cell = sheet.cell(row=row, column=whatsapp_col)
            cell.hyperlink = whatsapp_url
            cell.value = 'Send WhatsApp'
            cell.font = Font(color='0000FF', underline='single')
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # Save the formatted workbook
        book.save(output_file)
        print(f"Successfully created WhatsApp-enabled Excel file: {output_file}")
        
    except Exception as e:
        print(f"An error occurred: {e}")
        raise

if __name__ == "__main__":
    # Example usage
    input_file = input("Enter the path to your Excel file: ")
    create_whatsapp_template(input_file)
